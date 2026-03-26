"""
actualizar.py — Control de Vacunas Grupo 020
Genera: cache_data.json | cache_units.json | cache_vaccines.json

ALMACÉN OOAD:
  Es un stock único a nivel delegación. El sistema repite el mismo valor en
  cada fila de cada unidad solo como referencia. Se guarda UNA sola vez por
  vacuna en el JSON; el tablero lo muestra como dato global, no por zona.

Uso:    python actualizar.py
Req:    pip install openpyxl
"""
import sys, json, re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

SCRIPT_DIR   = Path(__file__).parent
DEFAULT_XLSX = SCRIPT_DIR / "Control_Existencias_por_Unidad.xlsx"
OUT_ZONES    = SCRIPT_DIR / "cache_data.json"
OUT_UNITS    = SCRIPT_DIR / "cache_units.json"
OUT_VACCINES = SCRIPT_DIR / "cache_vaccines.json"
DICC_XLSX    = SCRIPT_DIR / "dicc_unidades_coahuila.xlsx"

COL_UNIT = 5;  COL_DESC = 17;  COL_EX_U = 34;  COL_EX_A = 36

ZONES = ['CARBONÍFERA', 'CENTRO', 'LAGUNA', 'NORTE', 'SUR']

MANUAL_MAP = {
    'Direccion UMF 14':              'CARBONÍFERA',
    'Direccion UMF 15':              'LAGUNA',
    'Direccion UMF 26':              'LAGUNA',
    'Direccion UMF 61':              'CARBONÍFERA',
    'Direccion UMF 62':              'CARBONÍFERA',
    'Direccion UMF 64':              'CARBONÍFERA',
    'Direccion UMF 88':              'CENTRO',
    'Direccion UMR-EM 52':           'SUR',
    'Farmacia HGSZ-MF 13':          'CARBONÍFERA',
    'Farmacia HGSZ-MF 21':          'LAGUNA',
    'Farmacia HGSZ-MF 6':           'NORTE',
    'Farmacia HGZ 11':              'LAGUNA',
    'Farmacia HGZ-MF 24':           'LAGUNA',
    'Farmacia HRS Ramos':           'NORTE',
    'Farmacia HRS San Buenaventura':'NORTE',
}

UNIT_ZONE: dict = {}


def load_dict():
    if not DICC_XLSX.exists():
        print("  AVISO: dicc_unidades_coahuila.xlsx no encontrado → usando complemento manual.")
        UNIT_ZONE.update(MANUAL_MAP); return
    wb   = openpyxl.load_workbook(DICC_XLSX, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True)); wb.close()
    hdr  = [str(c).upper().strip() if c else '' for c in rows[0]]
    cu = cz = None
    for i, h in enumerate(hdr):
        if cu is None and 'NOMBRE' in h: cu = i
        if cz is None and 'ZONA'   in h: cz = i
    n = 0
    for row in rows[1:]:
        u = str(row[cu] or '').strip()
        z = str(row[cz] or '').strip().upper().replace('CARBONIFERA', 'CARBONÍFERA')
        if u and z in ZONES: UNIT_ZONE[u] = z; n += 1
    for u, z in MANUAL_MAP.items():
        if u not in UNIT_ZONE: UNIT_ZONE[u] = z
    print(f"  {n} del diccionario + {len(MANUAL_MAP)} del complemento manual.")


def get_zone(unit):
    if unit in UNIT_ZONE: return UNIT_ZONE[unit]
    m = re.search(r'\b(\d+)\b', unit)
    if m:
        for k, v in UNIT_ZONE.items():
            if re.search(rf'\b{m.group(1)}\b', k): return v
    return 'CENTRO'


def short_name(desc):
    d = str(desc).upper()
    if 'ANTIALACRAN'   in d: return 'ANTIALACRAN'
    if 'ANTIARACNIDO'  in d: return 'ANTIARACNIDO F/I'
    if 'ANTICORALILLO' in d: return 'ANTICORALILLO'
    if 'ANTIVIPERINO'  in d: return 'ANTIVIPERINO'
    if 'BCG'           in d: return 'BCG'
    if 'COVID' in d or ('ARNM' in d and 'SARS' in d): return 'COVID-19'
    if 'DOBLE VIRAL'   in d: return 'DOBLE VIRAL (SR)'
    if ('ANTIPERTUSSIS' in d and 'DIFTERICO' in d and 'TETANICO' in d
            and 'HEXAVALENTE' not in d and 'HEPATITIS' not in d): return 'DPT'
    if 'ANTITETANICA'  in d and 'INMUNOGLOBULINA' in d: return 'GAMA ANITETA...'
    if 'ANTIRRABICA'   in d and 'INMUNOGLOBULINA' in d: return 'GAMA ANITRA...'
    if 'HEPATITIS A'   in d and ('ADULTO' in d or '1.0 ML' in d): return 'HEPATITIS A AD...'
    if 'HEPATITIS A'   in d: return 'HEPATITIS A'
    if 'HEPATITIS B'   in d and 'DIFTERIA' not in d: return 'HEPATITIS B'
    if 'DIFTERIA'      in d and 'HEPATITIS B' in d: return 'HEXAVALENTE'
    if 'INFLUENZA'     in d: return 'INFLUENZA TET...'
    if '13-VALENTE'    in d: return 'NEUMO 13 V'
    if '20-VALENTE'    in d: return 'NEUMO 20 V'
    if 'ANTINEUMOCOCCICA' in d or 'NEUMOCOCICA' in d: return 'NEUMO 23 V'
    if 'ROTAVIRUS'     in d: return 'ROTAVIRUS'
    if 'TOXOIDES TETANICO Y DIFTERICO' in d: return 'TD'
    if 'TOSFERINA ACELULAR' in d: return 'TDPA'
    if 'TRIPLE VIRAL'  in d: return 'TRIPLE VIRAL (S...)'
    if 'PAPILOMA' in d or 'VPH' in d: return 'VHP 9 V'
    if 'VITAMINA A'    in d: return 'VITAMINA A'
    if 'ANTIRRABICA'   in d: return 'VACUNA ANTIR...'
    if 'VARICELA'      in d: return 'VARICELA'
    return d[:20]


def process(xlsx_path):
    print(f"  Leyendo: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_rows = list(wb.active.iter_rows(values_only=True)); wb.close()

    # Paso A — leer todas las filas
    # vac_almacen_ooad: valor único global del almacén OOAD por vacuna (mismo en todas las filas)
    vac_almacen_ooad: dict = {}
    # acumular existencias en unidad por (unidad, vacuna)
    unit_vac_unidad: dict = {}   # {unit: {vac: float}}
    unit_zone: dict = {}

    for row in all_rows[2:]:
        try:
            unit = str(row[COL_UNIT] or '').strip()
            desc = str(row[COL_DESC] or '').strip()
            ex_u = float(row[COL_EX_U] or 0)
            ex_a = float(row[COL_EX_A] or 0)
        except Exception: continue
        if not unit or not desc or unit == 'None': continue

        vac  = short_name(desc)
        zone = get_zone(unit)
        unit_zone[unit] = zone

        # almacén OOAD: guardar el valor una sola vez (es idéntico en todas las filas)
        if ex_a > 0 and vac not in vac_almacen_ooad:
            vac_almacen_ooad[vac] = int(ex_a)

        if unit not in unit_vac_unidad:
            unit_vac_unidad[unit] = {}
        unit_vac_unidad[unit][vac] = unit_vac_unidad[unit].get(vac, 0) + ex_u

    # Paso B — construir cache_units
    # Para cada unidad: existencias en unidad (reales) + almacén OOAD (referencia global)
    units_data: dict = {}
    for unit in sorted(unit_vac_unidad, key=lambda u: (
            ZONES.index(unit_zone.get(u, 'CENTRO')) if unit_zone.get(u,'CENTRO') in ZONES else 99, u)):
        zone = unit_zone.get(unit, 'CENTRO')
        vacs = {}
        for vac, ex_u in sorted(unit_vac_unidad[unit].items()):
            ex_a  = vac_almacen_ooad.get(vac, 0)
            vacs[vac] = {
                'unidad':  int(ex_u),
                'almacen': ex_a,          # referencia: stock OOAD global
                'total':   int(ex_u) + ex_a,
            }
        units_data[unit] = {'zona': zone, 'vacunas': vacs}

    # Paso C — construir cache_zones
    # unidad: suma real de existencias en unidad por zona
    # almacen: valor único del OOAD (NO multiplicar por zonas)
    zones_data = {z: {'vacunas': {}, 'unidades': []} for z in ZONES}
    for unit, ud in units_data.items():
        zone = ud['zona']
        if unit not in zones_data[zone]['unidades']:
            zones_data[zone]['unidades'].append(unit)
        for vac, vals in ud['vacunas'].items():
            if vac not in zones_data[zone]['vacunas']:
                zones_data[zone]['vacunas'][vac] = {'unidad': 0, 'almacen': 0, 'total': 0}
            zones_data[zone]['vacunas'][vac]['unidad'] += vals['unidad']
            # almacén: tomar una sola vez (idempotente: el valor es siempre el mismo)
            zones_data[zone]['vacunas'][vac]['almacen'] = vac_almacen_ooad.get(vac, 0)

    for z in ZONES:
        zones_data[z]['unidades'].sort()
        for vac in zones_data[z]['vacunas']:
            v = zones_data[z]['vacunas'][vac]
            v['total'] = v['unidad'] + v['almacen']
        zones_data[z]['vacunas'] = dict(sorted(zones_data[z]['vacunas'].items()))

    n = sum(len(v) for v in unit_vac_unidad.values())
    print(f"  Filas procesadas: {n} · Unidades: {len(units_data)}")
    return zones_data, units_data, vac_almacen_ooad


def build_vaccines(zones_data, units_data, vac_almacen_ooad):
    vaccines: dict = {}

    # zonas: existencias en unidad por zona + almacén global único
    for z, zd in zones_data.items():
        for vac, vals in zd['vacunas'].items():
            if vac not in vaccines:
                vaccines[vac] = {'almacen_ooad': vac_almacen_ooad.get(vac, 0),
                                  'zonas': {}, 'unidades': []}
            vaccines[vac]['zonas'][z] = {
                'unidad':  vals['unidad'],
                'almacen': vac_almacen_ooad.get(vac, 0),  # mismo valor para todas las zonas
                'total':   vals['unidad'] + vac_almacen_ooad.get(vac, 0),
            }

    # unidades
    for uname, ud in units_data.items():
        for vac, vals in ud['vacunas'].items():
            if vac not in vaccines:
                vaccines[vac] = {'almacen_ooad': vac_almacen_ooad.get(vac, 0),
                                  'zonas': {}, 'unidades': []}
            vaccines[vac]['unidades'].append({
                'nombre':  uname,
                'zona':    ud['zona'],
                'unidad':  vals['unidad'],
                'almacen': vac_almacen_ooad.get(vac, 0),
                'total':   vals['unidad'] + vac_almacen_ooad.get(vac, 0),
            })

    for vac in vaccines:
        vaccines[vac]['unidades'].sort(
            key=lambda x: (ZONES.index(x['zona']) if x['zona'] in ZONES else 99, x['nombre']))

    return dict(sorted(vaccines.items()))


if __name__ == '__main__':
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"\nERROR: No se encontró {xlsx_path}"); sys.exit(1)

    print("① Cargando diccionario de zonas...")
    load_dict()
    print("② Procesando existencias...")
    zones_data, units_data, vac_almacen_ooad = process(xlsx_path)
    print("③ Generando índice por vacuna...")
    vaccines_data = build_vaccines(zones_data, units_data, vac_almacen_ooad)
    print("④ Guardando archivos...")
    with open(OUT_ZONES,    'w', encoding='utf-8') as f: json.dump(zones_data,    f, ensure_ascii=False, indent=2)
    with open(OUT_UNITS,    'w', encoding='utf-8') as f: json.dump(units_data,    f, ensure_ascii=False, indent=2)
    with open(OUT_VACCINES, 'w', encoding='utf-8') as f: json.dump(vaccines_data, f, ensure_ascii=False, indent=2)

    print(f"\n✅  Tres archivos generados:")
    print(f"    {OUT_ZONES}\n    {OUT_UNITS}\n    {OUT_VACCINES}")

    print("\nResumen por zona:")
    print(f"  {'ZONA':<14} {'UMED':>5}  {'VAC':>4}  {'EN UNIDAD':>10}  {'ALMACÉN OOAD':>13}  {'TOTAL':>10}")
    print(f"  {'-'*14} {'-'*5}  {'-'*4}  {'-'*10}  {'-'*13}  {'-'*10}")
    for z in ZONES:
        tu = sum(v['unidad']  for v in zones_data[z]['vacunas'].values())
        ta = sum(v['almacen'] for v in zones_data[z]['vacunas'].values())
        tt = sum(v['total']   for v in zones_data[z]['vacunas'].values())
        print(f"  {z:<14} {len(zones_data[z]['unidades']):>5}  {len(zones_data[z]['vacunas']):>4}  {tu:>10,}  {ta:>13,}  {tt:>10,}")

    print("\nVerificación HEPATITIS A — almacén debe ser 8 (no multiplicado):")
    for z in ZONES:
        v = zones_data[z]['vacunas'].get('HEPATITIS A')
        if v: print(f"  {z:<14} unidad={v['unidad']:>5,}  almacén={v['almacen']:>4,}  total={v['total']:>5,}")
    print(f"  almacén_ooad global = {vac_almacen_ooad.get('HEPATITIS A', 0)}")
    print("\n→ Recarga el tablero en tu navegador.")
