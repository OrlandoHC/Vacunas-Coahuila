"""
actualizar_remisiones.py — Remisiones Semanales Grupo 020
Genera: rem_zones.json | rem_units.json | rem_vacs.json | rem_meta.json

Uso:    python actualizar_remisiones.py
Req:    pip install openpyxl
"""
import sys, json, re
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

SCRIPT_DIR   = Path(__file__).parent
DEFAULT_XLSX = SCRIPT_DIR / "Remisiones_Delegacion.xlsx"
DICC_XLSX    = SCRIPT_DIR / "dicc_unidades_coahuila.xlsx"

ZONES = ['CARBONÍFERA','CENTRO','LAGUNA','NORTE','SUR']

MANUAL_MAP = {
    'Direccion UMF 14':'CARBONÍFERA','Direccion UMF 15':'LAGUNA',
    'Direccion UMF 26':'LAGUNA','Direccion UMF 60':'SUR',
    'Direccion UMF 61':'CARBONÍFERA','Direccion UMF 62':'CARBONÍFERA',
    'Direccion UMF 64':'CARBONÍFERA','Direccion UMF 88':'CENTRO',
    'Direccion UMR-EM 52':'SUR','Farmacia HGSZ-MF 13':'CARBONÍFERA',
    'Farmacia HGSZ-MF 21':'LAGUNA','Farmacia HGSZ-MF 6':'NORTE',
    'Farmacia HGZ 11':'LAGUNA','Farmacia HGZ-MF 24':'LAGUNA',
    'Farmacia HRS Ramos':'NORTE','Farmacia HRS San Buenaventura':'NORTE',
    'U Med Familiar 50 Direccion de la Unidad Médica':'CENTRO',
    'U Med Familiar 60 Direccion de la Unidad Médica':'SUR',
    'U Med Familiar 62 Direccion de la Unidad Médica':'CARBONÍFERA',
    'U Med Familiar 64 Direccion de la Unidad Médica':'CARBONÍFERA',
}

UNIT_ZONE: dict = {}

def load_dict():
    if not DICC_XLSX.exists():
        print("  AVISO: dicc_unidades_coahuila.xlsx no encontrado → usando complemento manual.")
        UNIT_ZONE.update(MANUAL_MAP); return
    wb   = openpyxl.load_workbook(DICC_XLSX, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True)); wb.close()
    hdr  = [str(c).upper().strip() if c else '' for c in rows[0]]
    cu   = next((i for i,h in enumerate(hdr) if 'NOMBRE' in h), None)
    cz   = next((i for i,h in enumerate(hdr) if 'ZONA'   in h), None)
    n = 0
    for row in rows[1:]:
        u = str(row[cu] or '').strip()
        z = str(row[cz] or '').strip().upper().replace('CARBONIFERA','CARBONÍFERA')
        if u and z in ZONES: UNIT_ZONE[u] = z; n += 1
    for u,z in MANUAL_MAP.items():
        if u not in UNIT_ZONE: UNIT_ZONE[u] = z
    print(f"  {n} del diccionario + {len(MANUAL_MAP)} del complemento manual.")

def get_zone(unit):
    if unit in UNIT_ZONE: return UNIT_ZONE[unit]
    m = re.search(r'\b(\d+)\b', unit)
    if m:
        for k,v in UNIT_ZONE.items():
            if re.search(rf'\b{m.group(1)}\b', k): return v
    return 'CENTRO'

def short_name(desc):
    d = str(desc).upper()
    if 'ANTIALACRAN'   in d: return 'ANTIALACRAN'
    if 'ANTIARACNIDO'  in d: return 'ANTIARACNIDO F/I'
    if 'ANTICORALILLO' in d: return 'ANTICORALILLO'
    if 'ANTIVIPERINO'  in d: return 'ANTIVIPERINO'
    if 'BCG'           in d: return 'BCG'
    if 'COVID' in d or ('ARNM' in d and 'SARS' in d): return 'COVID-19'
    if 'DOBLE VIRAL'   in d: return 'DOBLE VIRAL (SR)'
    if ('ANTIPERTUSSIS' in d and 'DIFTERICO' in d and 'TETANICO' in d
            and 'HEXAVALENTE' not in d and 'HEPATITIS' not in d): return 'DPT'
    if 'ANTITETANICA'  in d and 'INMUNOGLOBULINA' in d: return 'GAMA ANITETA...'
    if 'ANTIRRABICA'   in d and 'INMUNOGLOBULINA' in d: return 'GAMA ANITRA...'
    if 'HEPATITIS A'   in d and ('ADULTO' in d or '1.0 ML' in d): return 'HEPATITIS A AD...'
    if 'HEPATITIS A'   in d: return 'HEPATITIS A'
    if 'HEPATITIS B'   in d and 'DIFTERIA' not in d: return 'HEPATITIS B'
    if 'DIFTERIA'      in d and 'HEPATITIS B' in d: return 'HEXAVALENTE'
    if 'INFLUENZA'     in d: return 'INFLUENZA TET...'
    if '13-VALENTE'    in d: return 'NEUMO 13 V'
    if '20-VALENTE'    in d: return 'NEUMO 20 V'
    if 'ANTINEUMOCOCCICA' in d or 'NEUMOCOCICA' in d: return 'NEUMO 23 V'
    if 'ROTAVIRUS'     in d: return 'ROTAVIRUS'
    if 'TOXOIDES TETANICO Y DIFTERICO' in d: return 'TD'
    if 'TOSFERINA ACELULAR' in d: return 'TDPA'
    if 'TRIPLE VIRAL'  in d: return 'TRIPLE VIRAL (S...)'
    if 'PAPILOMA' in d or 'VPH' in d: return 'VHP 9 V'
    if 'VITAMINA A'    in d: return 'VITAMINA A'
    if 'ANTIRRABICA'   in d: return 'VACUNA ANTIR...'
    if 'VARICELA'      in d: return 'VARICELA'
    return d[:20]

def rnd(d):
    return {k: int(v) if isinstance(v, float) else v for k,v in d.items()}

def process(xlsx_path):
    print(f"  Leyendo: {xlsx_path.name}")
    wb   = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True)); wb.close()

    zones_out = {z:{'vacunas':{},'unidades':[]} for z in ZONES}
    units_out: dict = {}
    vacs_out:  dict = {}
    fechas = set()

    for row in rows[10:]:   # datos desde fila 11 (índice 10)
        unit  = str(row[5]  or '').strip()
        desc  = str(row[20] or '').strip()
        fecha = row[10]
        env   = float(row[22] or 0)
        rec   = float(row[23] or 0)
        dev   = float(row[24] or 0)
        imp   = float(row[25] or 0)
        if not unit or not desc: continue

        zone = get_zone(unit)
        vac  = short_name(desc)
        fstr = str(fecha)[:10] if fecha else 'Sin fecha'
        fechas.add(fstr)

        # — zona —
        if unit not in zones_out[zone]['unidades']:
            zones_out[zone]['unidades'].append(unit)
        zv = zones_out[zone]['vacunas']
        if vac not in zv: zv[vac] = {'enviada':0,'recibida':0,'devuelta':0,'importe':0}
        zv[vac]['enviada']  += env; zv[vac]['recibida'] += rec
        zv[vac]['devuelta'] += dev; zv[vac]['importe']  += imp

        # — unidad —
        if unit not in units_out: units_out[unit] = {'zona':zone,'vacunas':{}}
        uv = units_out[unit]['vacunas']
        if vac not in uv: uv[vac] = {'enviada':0,'recibida':0,'devuelta':0,'importe':0}
        uv[vac]['enviada']  += env; uv[vac]['recibida'] += rec
        uv[vac]['devuelta'] += dev; uv[vac]['importe']  += imp

        # — vacuna —
        if vac not in vacs_out: vacs_out[vac] = {'zonas':{},'unidades':[]}
        if zone not in vacs_out[vac]['zonas']:
            vacs_out[vac]['zonas'][zone] = {'enviada':0,'recibida':0,'devuelta':0,'importe':0}
        vv = vacs_out[vac]['zonas'][zone]
        vv['enviada']  += env; vv['recibida'] += rec
        vv['devuelta'] += dev; vv['importe']  += imp

    # Ordenar y redondear
    for z in ZONES:
        zones_out[z]['unidades'].sort()
        zones_out[z]['vacunas'] = {k:rnd(v) for k,v in sorted(zones_out[z]['vacunas'].items())}

    units_out = dict(sorted(units_out.items(),
        key=lambda x:(ZONES.index(x[1]['zona']) if x[1]['zona'] in ZONES else 99, x[0])))
    for u in units_out:
        units_out[u]['vacunas'] = {k:rnd(v) for k,v in sorted(units_out[u]['vacunas'].items())}

    for v in vacs_out:
        vacs_out[v]['zonas'] = {k:rnd(vv) for k,vv in vacs_out[v]['zonas'].items()}
        vacs_out[v]['unidades'] = sorted([
            {'nombre':u,'zona':ud['zona'],**rnd(ud['vacunas'][v])}
            for u,ud in units_out.items() if v in ud['vacunas']
        ], key=lambda x:(ZONES.index(x['zona']) if x['zona'] in ZONES else 99, x['nombre']))
    vacs_out = dict(sorted(vacs_out.items()))

    fechas_s = sorted(f for f in fechas if f != 'Sin fecha')
    meta = {
        'fecha_inicio':   fechas_s[0] if fechas_s else '',
        'fecha_fin':      fechas_s[-1] if fechas_s else '',
        'total_piezas':   int(sum(sum(v['enviada'] for v in ud['vacunas'].values()) for ud in units_out.values())),
        'total_recibidas':int(sum(sum(v['recibida'] for v in ud['vacunas'].values()) for ud in units_out.values())),
        'total_importe':  round(sum(sum(v['importe'] for v in ud['vacunas'].values()) for ud in units_out.values()), 2),
        'total_unidades': len(units_out),
        'total_vacunas':  len(vacs_out),
    }
    print(f"  Filas procesadas: {len(rows)-10} · Unidades: {len(units_out)} · Vacunas: {len(vacs_out)}")
    return zones_out, units_out, vacs_out, meta

if __name__ == '__main__':
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"\nERROR: No se encontró {xlsx_path}")
        print("Coloca 'Remisiones_Delegacion.xlsx' en esta carpeta.\n"); sys.exit(1)

    print("① Cargando diccionario de zonas...")
    load_dict()
    print("② Procesando remisiones...")
    zones_data, units_data, vacs_data, meta = process(xlsx_path)

    print("③ Guardando archivos...")
    (SCRIPT_DIR/'rem_zones.json').write_text(json.dumps(zones_data, ensure_ascii=False, indent=2))
    (SCRIPT_DIR/'rem_units.json').write_text(json.dumps(units_data, ensure_ascii=False, indent=2))
    (SCRIPT_DIR/'rem_vacs.json' ).write_text(json.dumps(vacs_data,  ensure_ascii=False, indent=2))
    (SCRIPT_DIR/'rem_meta.json' ).write_text(json.dumps(meta,       ensure_ascii=False, indent=2))

    print(f"\n✅  Cuatro archivos generados:")
    for f in ['rem_zones.json','rem_units.json','rem_vacs.json','rem_meta.json']:
        print(f"    {SCRIPT_DIR/f}")

    print(f"\nPeriodo: {meta['fecha_inicio']} → {meta['fecha_fin']}")
    print(f"Piezas enviadas: {meta['total_piezas']:,}   Importe: ${meta['total_importe']:,.2f}")
    print(f"\nResumen por zona:")
    print(f"  {'ZONA':<14} {'UMED':>5}  {'VAC':>4}  {'ENVIADAS':>10}  {'RECIBIDAS':>10}  {'IMPORTE':>15}")
    print(f"  {'-'*14} {'-'*5}  {'-'*4}  {'-'*10}  {'-'*10}  {'-'*15}")
    for z in ZONES:
        te = sum(v['enviada']  for v in zones_data[z]['vacunas'].values())
        tr = sum(v['recibida'] for v in zones_data[z]['vacunas'].values())
        ti = sum(v['importe']  for v in zones_data[z]['vacunas'].values())
        print(f"  {z:<14} {len(zones_data[z]['unidades']):>5}  "
              f"{len(zones_data[z]['vacunas']):>4}  {te:>10,}  {tr:>10,}  ${ti:>14,.2f}")
    print("\n→ Recarga el tablero en tu navegador.")
